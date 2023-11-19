from genericpath import isfile
# from os import *
import tkinter as tk
import openpyxl
window = tk.Tk()
window.geometry('600x600')
window.config(background="grey")



def save_details():
    p_name = input1.get()
    p_description = input2.get()
    p_price = input3.get()
    
    
    if isfile('untitled1.xlsx'):
        pass
    #saving on asn external excel
    else:
     wb = openpyxl.Workbook()
     ws = wb.create_sheet("sheet1")
     ws.append(['p-name','p_description','p_price'])
     wb.save("untitled1.xlsx")
    
    wb = openpyxl.load_workbook("untitled1.xlsx")
    ws = wb.worksheets[0]
    ws.append([p_name,p_description,p_price])
    wb.save("untitled1.xlsx")
    
    
    cell = ws["A1"]
    value = cell.value
    print(value)
    

    
    



def dialog_box():
    dialog = tk.Toplevel()
    dialog.geometry("400x400")
    
    label = tk.Label(dialog,text="are you sure ?")
    label.pack()
    
    yes = tk.Button(dialog,text="yes",command=save_details)
    yes.pack()
    
    no = tk.Button(dialog,text="no")
    no.pack()


label = tk.Label(window,text="prod_name",pady="20",background="white")
label.config(background="red",width="100")
label.pack()

input1 = tk.Entry(window)
input1.config(width="100")
input1.pack()

label = tk.Label(window,text="prod_description",pady="20")
label.config(background="red",width="100")
label.pack()

input2 = tk.Entry(window)
input2.config(width="100")
input2.pack()

label = tk.Label(window,text="prod_price",pady="20")
label.config(background="red",width="100")
label.pack()

input3 = tk.Entry(window)
input3.config(width="100")
input3.pack()


save_button = tk.Button(window,text="submit",command=dialog_box)
save_button.pack()


window.mainloop()
